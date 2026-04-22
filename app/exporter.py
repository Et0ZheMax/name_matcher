from __future__ import annotations

import csv
from pathlib import Path

from app.models import PipelineResult


def export_result(result: PipelineResult, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        _write_csv(result, output_path)
        return

    try:
        from openpyxl import Workbook
    except Exception:
        fallback = output_path.with_suffix(".csv")
        _write_csv(result, fallback)
        return

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "organizations_enriched"
    _write_sheet(ws1, _org_rows(result))

    ws2 = wb.create_sheet("original_plus_enrichment")
    _write_sheet(ws2, result.original_rows_enriched)

    ws3 = wb.create_sheet("manual_review")
    _write_sheet(
        ws3,
        [
            {
                "organization_ru_raw": x.organization_ru_raw,
                "organization_en_final": x.organization_en_final,
                "final_status": x.final_status,
                "final_confidence": x.final_confidence,
                "source_primary": x.source_primary,
                "source_secondary": x.source_secondary,
                "pubmed_status": x.pubmed.pubmed_validation_status,
                "pubmed_exact_count": x.pubmed.pubmed_exact_count,
                "notes": x.notes,
            }
            for x in result.manual_review
        ],
    )

    ws4 = wb.create_sheet("candidates_debug")
    _write_sheet(
        ws4,
        [
            {
                "organization_ru_raw": c.organization_ru_raw,
                "organization_ru_normalized": c.organization_ru_normalized,
                "candidate_text": c.candidate_text,
                "normalized_candidate_text": c.normalized_candidate_text,
                "source": c.source,
                "contributing_sources": ", ".join(c.contributing_sources),
                "source_url": c.source_url,
                "support_signals": ", ".join(c.support_signals),
                "score": c.score,
                "confidence": c.confidence,
                "source_evidence": " | ".join(f"{e.get('source')}: {e.get('text')}" for e in c.source_evidence),
                "notes": "; ".join(c.notes),
            }
            for c in result.candidates_debug
        ],
    )

    wb.save(output_path)


def _write_csv(result: PipelineResult, output_path: Path) -> None:
    rows = _org_rows(result)
    if not rows:
        return
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _org_rows(result: PipelineResult):
    return [
        {
            "organization_ru_raw": x.organization_ru_raw,
            "organization_ru_normalized": x.organization_ru_normalized,
            "organization_en_final": x.organization_en_final,
            "final_status": x.final_status,
            "final_confidence": x.final_confidence,
            "source_primary": x.source_primary,
            "source_url_primary": x.source_url_primary,
            "source_secondary": x.source_secondary,
            "website_url": x.website_url,
            "ror_id": x.ror_id,
            "wikipedia_url": x.wikipedia_url,
            "wikidata_url": x.wikidata_url,
            "pubmed_exact_query": x.pubmed.pubmed_exact_query,
            "pubmed_exact_count": x.pubmed.pubmed_exact_count,
            "pubmed_broad_query": x.pubmed.pubmed_broad_query,
            "pubmed_broad_count": x.pubmed.pubmed_broad_count,
            "pubmed_validation_status": x.pubmed.pubmed_validation_status,
            "pubmed_validation_notes": x.pubmed.pubmed_validation_notes,
            "pubmed_pmids": ",".join(x.pubmed.pubmed_pmids),
            "notes": x.notes,
        }
        for x in result.organizations
    ]


def _write_sheet(ws, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
