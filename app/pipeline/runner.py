from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Callable, Dict, List, Optional

from app.config import AppConfig
from app.models import PipelineResult, PubmedValidation
from app.pipeline.candidate_builder import CandidateBuilder
from app.pipeline.normalize import normalize_org_name
from app.pipeline.resolver import Resolver
from app.pipeline.validator import CandidateValidator


class PipelineRunner:
    def __init__(
        self,
        config: AppConfig,
        candidate_builder: CandidateBuilder,
        resolver: Resolver,
        validator: CandidateValidator,
        logger: logging.Logger,
    ) -> None:
        self.config = config
        self.builder = candidate_builder
        self.resolver = resolver
        self.validator = validator
        self.logger = logger

    def run(
        self,
        input_path: Path,
        org_column: Optional[str] = None,
        first_column_as_org: bool = False,
        limit: Optional[int] = None,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> PipelineResult:
        rows = self._read_input(input_path)
        if limit:
            rows = rows[:limit]
        if not rows:
            return PipelineResult([], [], [], [])

        org_col = self._detect_column(rows[0], org_column, first_column_as_org)

        normalized_map: Dict[str, object] = {}
        for row in rows:
            n = normalize_org_name(str(row.get(org_col, "")))
            normalized_map.setdefault(n.normalized, n)

        unique_orgs = list(normalized_map.values())
        enriched = []
        candidates_debug = []

        for idx, org in enumerate(unique_orgs, start=1):
            self.logger.info("Processing [%d/%d]: %s", idx, len(unique_orgs), org.raw)
            if progress_callback:
                progress_callback(idx, len(unique_orgs), org.raw)

            candidates = self.builder.build(org)
            _best, _status, ranked = self.resolver.resolve(candidates)

            pubmed_by_candidate: Dict[int, PubmedValidation] = {}
            for candidate in ranked[: min(3, len(ranked))]:
                validated, pub = self.validator.validate_with_pubmed(candidate)
                pubmed_by_candidate[id(validated)] = pub

            best, status, ranked = self.resolver.resolve(ranked)
            pub = pubmed_by_candidate.get(id(best), PubmedValidation())

            resolved = self.resolver.build_resolved(best, status)
            resolved.pubmed = pub
            enriched.append(resolved)
            candidates_debug.extend(ranked)
            if status == "manual_review_needed" or resolved.final_confidence < self.config.manual_review_confidence_threshold:
                self.logger.warning(
                    "Manual review reason: status=%s confidence=%.2f candidate=%s",
                    status,
                    resolved.final_confidence,
                    resolved.organization_en_final,
                )

        by_norm = {x.organization_ru_normalized: x for x in enriched}
        original_plus = []
        for row in rows:
            n = normalize_org_name(str(row.get(org_col, "")))
            match = by_norm.get(n.normalized)
            merged = dict(row)
            if match:
                merged.update(
                    {
                        "organization_ru_normalized": match.organization_ru_normalized,
                        "organization_en_final": match.organization_en_final,
                        "final_status": match.final_status,
                        "final_confidence": match.final_confidence,
                    }
                )
            original_plus.append(merged)

        manual = [
            x
            for x in enriched
            if x.final_status == "manual_review_needed"
            or x.final_confidence < self.config.manual_review_confidence_threshold
        ]
        return PipelineResult(enriched, original_plus, manual, candidates_debug)

    @staticmethod
    def _read_input(input_path: Path) -> List[Dict[str, str]]:
        suffix = input_path.suffix.lower()
        if suffix in {".xlsx", ".xls"}:
            try:
                from openpyxl import load_workbook
            except Exception as exc:
                raise RuntimeError("openpyxl is required for Excel input") from exc
            wb = load_workbook(input_path, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h) for h in rows[0]]
            return [dict(zip(headers, ["" if v is None else str(v) for v in r])) for r in rows[1:]]

        with input_path.open("r", encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))

    @staticmethod
    def _detect_column(first_row: Dict[str, str], org_column: Optional[str], first_column_as_org: bool) -> str:
        cols = list(first_row.keys())
        if org_column:
            if org_column not in cols:
                raise ValueError(f"Column '{org_column}' not found")
            return org_column
        if first_column_as_org or cols:
            return cols[0]
        raise ValueError("Could not detect organization column")
